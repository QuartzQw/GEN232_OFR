import os
import pickle
from tkinter import messagebox
import pandas as pd
import cv2
import numpy as np
from PIL import Image as PILImage, Image
from openpyxl.drawing.image import Image as opxImage
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from tensorflow.keras.models import load_model
from uuid import uuid4
from datetime import datetime
import re

MODEL_DIGIT = None

def to_excel_col(num):
    col = ''
    while num:
        num, rem = divmod(num - 1, 26)
        col = chr(65 + rem) + col
    return col

def preprocess_image(cropped_image):
    img = cropped_image.convert('L')
    img = np.array(img)
    img = cv2.bitwise_not(img)
    _, thresh = cv2.threshold(img, 82, 255, cv2.THRESH_BINARY)
    return thresh

def find_digits(image):
    contours, _ = cv2.findContours(image.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    # digit_rects = sorted([cv2.boundingRect(c) for c in contours], key=lambda x: x[0])
    rects = []
    for c in contours:
        if cv2.contourArea(c) > 7:
            rects.append(c)
    digit_rects = sorted([cv2.boundingRect(c) for c in rects], key=lambda x: x[0])
    return digit_rects

def extract_digits(image, digit_rects, padding=5):
    digits = []
    for x, y, w, h in digit_rects:
        digit = image[max(0, y - padding):y + h + padding, max(0, x - padding):x + w + padding]
        digit = cv2.resize(digit, (28, 28)).astype('float32') / 255.0
        digit = np.expand_dims(digit, axis=-1)
        digits.append(digit)
    return np.array(digits)

import sys

def classify_handwritten_digits(cropped_image):
    global MODEL_DIGIT
    if MODEL_DIGIT is None:
        # ใช้ path สำหรับ pyinstaller (exe)
        base_dir = getattr(sys, '_MEIPASS', os.path.abspath("."))
        model_path = os.path.join(base_dir, 'mnist_cnn.h5')
        MODEL_DIGIT = load_model(model_path)
    
    processed = preprocess_image(cropped_image)
    boxes = find_digits(processed)
    if not boxes:
        return ""   #[no digits]
    digits = extract_digits(processed, boxes)
    pred = MODEL_DIGIT.predict(digits)
    return ''.join(str(np.argmax(p)) for p in pred)

def extract_number_from_image(cropped_image):
    try:
        return classify_handwritten_digits(cropped_image)
    except Exception as e:
        return f"[error: {e}]"

def is_checkbox_checked(cropped_image):
    img_np = np.array(cropped_image.convert("L"))
    _, binary = cv2.threshold(img_np, 145, 255, cv2.THRESH_BINARY_INV)
    black_pixels = np.sum(binary == 255)
    return 1 if black_pixels > 0.45 * binary.size else 0

def crop_and_save_image(image, coord, folder, name, index):
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, f"{name}_{index}.png")
    image.save(path)
    return path

def get_image_and_coords(index, image_paths, template_pages):
    image = Image.open(image_paths[index])
    coords = template_pages[index % len(template_pages)]
    return image, coords

def extract_number(filename):
    match = re.search(r'(\d+)', filename)
    return int(match.group()) if match else -1

def process_survey(image_folder, template_file, output_file, offset_map=None, show_alert=True, append_excel_path=None):
    PERMANENT_IMAGE_DIR = "./saved_images"
    os.makedirs(PERMANENT_IMAGE_DIR, exist_ok=True)

    if offset_map is None:
        offset_map = {}

    temp_img_folder = f"./tempArea/images-{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    os.makedirs(temp_img_folder, exist_ok=True)

    with open(template_file, "rb") as f:
        data = pickle.load(f)
        template_pages = data["pages"]  # ไม่ reversed

    image_paths = [os.path.join(image_folder, f) for f in os.listdir(image_folder)
                   if f.lower().endswith(('.jpg', '.png', '.jpeg'))]
    image_paths.sort(key=lambda x: extract_number(os.path.basename(x)))

    field_order = []
    field_type_map = {}
    for page in template_pages:
        for coord in page:
            fname = str(coord[5])
            if fname not in field_order:
                field_order.append(fname)
            field_type_map[fname] = coord[4]

    results = []
    pages_per_doc = len(template_pages)

    for start in range(0, len(image_paths), pages_per_doc):
        batch_result = {}
        for page_idx in range(pages_per_doc):
            img_idx = start + page_idx
            if img_idx >= len(image_paths): break
            image, coords = get_image_and_coords(img_idx, image_paths, template_pages)
            w, h = image.size
            offset_x, offset_y = offset_map.get(page_idx, (0, 0))
            for coord in coords:
                try:
                    x1 = int(coord[0] * w) + offset_x
                    y1 = int(coord[1] * h) + offset_y
                    x2 = int(coord[2] * w) + offset_x
                    y2 = int(coord[3] * h) + offset_y
                    crop = image.crop((x1, y1, x2, y2))
                    ftype, fname = coord[4], str(coord[5])
                    if ftype == "checkBox":
                        batch_result[fname] = is_checkbox_checked(crop)
                    elif ftype == "image":
                        unique = uuid4().hex[:6]
                        path = os.path.join(temp_img_folder, f"page{page_idx}_{fname}_{unique}.jpg")
                        crop.save(path)
                        batch_result[fname] = path.replace("\\", "/")
                    elif ftype == "number":
                        batch_result[fname] = extract_number_from_image(crop)
                    elif ftype == "imageLink":
                        save_path = os.path.join(PERMANENT_IMAGE_DIR, f"{fname}_{uuid4().hex[:6]}.png")
                        crop.save(save_path)
                        batch_result[fname] = save_path.replace("\\", "/")
                except Exception as e:
                    print(f"[ERROR] Page {page_idx}, Field {coord}: {e}")
        results.append(batch_result)

    df_new = pd.DataFrame([{k: r.get(k, "") for k in field_order} for r in results])

    if append_excel_path and os.path.exists(append_excel_path):
        wb = load_workbook(append_excel_path)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        start_row = 2
        for col_idx, col_name in enumerate(df_new.columns, 1):
            ws[f"{get_column_letter(col_idx)}1"] = col_name

    for row_offset, row in enumerate(df_new.itertuples(index=False), start=start_row):
        for col_idx, val in enumerate(row, 1):
            cell = f"{get_column_letter(col_idx)}{row_offset}"
            if isinstance(val, str) and os.path.exists(val) and val.lower().endswith(('.jpg', '.png')):
                try:
                    col_name = df_new.columns[col_idx - 1]
                    field_type = field_type_map.get(col_name, "")

                    if field_type == "imageLink":
                        filename = os.path.basename(val)
                        relative_path = os.path.relpath(val, start=os.path.dirname(output_file))
                        ws[cell].value = filename
                        ws[cell].hyperlink = relative_path.replace(os.sep, '/')
                        ws[cell].style = 'Hyperlink'  # เปลี่ยนเป็นสีฟ้า
                    else:
                        img = opxImage(val)
                        pil_img = PILImage.open(val)
                        img.height = 18
                        img.width = int(pil_img.width / pil_img.height * 18)
                        img.anchor = cell
                        ws[cell].value = " "
                        ws.add_image(img)
                        ws.column_dimensions[get_column_letter(col_idx)].width = round((img.width + 3) / 7)

                except Exception as e:
                    print(f"Embed/link fail at {cell}: {e}")
            else:
                ws[cell] = val


    wb.save(output_file)

    for file in os.listdir(temp_img_folder):
        try:
            os.remove(os.path.join(temp_img_folder, file))
        except:
            pass
    try:
        os.rmdir(temp_img_folder)
    except:
        pass

    if show_alert:
        messagebox.showinfo("Export Completed", f"Excel file has been saved to:\n{output_file}")